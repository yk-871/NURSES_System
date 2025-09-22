import os
import pandas as pd
import joblib
import json
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from ortools.linear_solver import pywraplp
import random



# ---------------- Config ----------------
DATASET_CSV = "dataset/covid_dataset.csv"
MODELS_DIR = "models"
NURSE_JSON = "csv/nurse_database.json"  # Nurse database JSON
OUTPUT_EXCEL = "output_schedule.xlsx"

wards = ["ED", "GW", "ICU"]
shifts = ["Morning", "Evening", "Night"]
shift_indices = {s: i for i, s in enumerate(shifts)}  # Map shifts to indices

SHIFT_HOURS = 8
MIN_WEEKLY_HOURS = 40
MAX_WEEKLY_HOURS = 60

# Excel colors
fill_on_duty = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # green
fill_off = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")    # grey

# ---------------- Load Models ----------------
def load_models():
    models = {}
    for ward in wards:
        path = os.path.join(MODELS_DIR, f"{ward}_nurse_demand.pkl")
        try:
            models[ward] = joblib.load(path)
        except FileNotFoundError:
            raise FileNotFoundError(f"Model file not found: {path}")
        except Exception as e:
            raise Exception(f"Error loading model {ward}: {str(e)}")
    return models

# ---------------- Prediction ----------------
def predict_next_week(df, models, days=7):
    last_row = df.iloc[-1]
    feature_cols = ["New case", "ICU", "Admission"]
    X_last = pd.DataFrame([last_row[feature_cols].values], columns=feature_cols)
    last_date = pd.to_datetime(last_row["Date"])

    preds = []
    for i in range(1, days + 1):
        future_date = last_date + timedelta(days=i)
        day_pred = {"Date": future_date}
        for ward in wards:
            y = int(round(models[ward].predict(X_last)[0]))
            day_pred[ward] = max(2, y)  # at least 2 nurses always
        preds.append(day_pred)

    return pd.DataFrame(preds)

# ---------------- Scheduling with Optimization ----------------
def get_next_week_start():
    """Get the start date for next week, advancing by 7 days each generation"""
    # Check if there's a saved current week date
    week_file = 'current_week.txt'
    
    if os.path.exists(week_file):
        with open(week_file, 'r') as f:
            current_week_str = f.read().strip()
            current_week = datetime.strptime(current_week_str, '%Y-%m-%d').date()
            next_week_start = current_week + timedelta(days=7)
    else:
        # First time - start from Monday Sept 22, 2025
        next_week_start = datetime(2025, 9, 22).date()
    
    # Save the new week date
    with open(week_file, 'w') as f:
        f.write(next_week_start.strftime('%Y-%m-%d'))
    
    print(f"Next week starts: {next_week_start}")
    return next_week_start

def get_next_monday():
    """Get the date of next Monday"""
    current_monday = get_current_monday()
    return current_monday + timedelta(days=7)

def schedule_nurses_optimized(week_demand, nurses, wards, shifts, start_date=None, weeks=1):
    if start_date is None:
        start_date = get_next_week_start()
    
    total_days = 7 * weeks  # Generate for one week only
    
    total_shifts_per_day = len(shifts)
    
    # Generate dates for multiple weeks
    all_dates = [start_date + timedelta(days=i) for i in range(total_days)]
    day_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    
    print(f"Generating schedule from {start_date} for {total_days} days")
    print(f"Week dates: {[d.strftime('%Y-%m-%d') for d in all_dates[:14]]}")

    # Use nurse ID as key to prevent missing/duplicate issues
    nurse_ids = [n["id"] for n in nurses]

    # Create solver
    solver = pywraplp.Solver.CreateSolver('SCIP')
    if not solver:
        return None, None, None

    # Decision variables
    x = {}
    for n_idx in range(len(nurses)):
        for d in range(total_days):
            for s in range(total_shifts_per_day):
                x[n_idx, d, s] = solver.BoolVar(f'x[{n_idx},{d},{s}]')

    # 1. One shift per nurse per day
    for n_idx in range(len(nurses)):
        for d in range(total_days):
            solver.Add(sum(x[(n_idx, d, s)] for s in range(total_shifts_per_day)) <= 1)

    # 2. No back-to-back Night to Morning shifts
    for n_idx in range(len(nurses)):
        for d in range(total_days - 1):
            solver.Add(x[n_idx, d, 2] + x[n_idx, d + 1, 0] <= 1)

    # 3. Minimum 2 nurses per ward per shift
    for d in range(total_days):
        for w_idx, w in enumerate(wards):
            for s in range(total_shifts_per_day):
                eligible_nurses = [n_idx for n_idx, n in enumerate(nurses) 
                                   if w in n.get("skills", []) or 
                                   any(f"{w} {role}" in n.get("skills", []) 
                                       for role in ["Nurse", "Specialist", "Charge Nurse", "Nursing Officer", "Senior Staff Nurse"])]
                solver.Add(sum(x[n_idx, d, s] for n_idx in eligible_nurses) >= 2)

    # 4. Max shifts per week and weekly hours
    for n_idx, nurse in enumerate(nurses):
        max_shifts = nurse.get("max_shifts_per_week", 7)
        solver.Add(sum(x[n_idx, d, s] for d in range(total_days) for s in range(total_shifts_per_day)) <= max_shifts)
        total_hours = sum(x[n_idx, d, s] * SHIFT_HOURS for d in range(total_days) for s in range(total_shifts_per_day))
        solver.Add(total_hours <= MAX_WEEKLY_HOURS)

    # Objective: Minimize unused capacity
    solver.Minimize(sum(1 - sum(x[n_idx, d, s] for d in range(total_days) for s in range(total_shifts_per_day)) 
                        for n_idx in range(len(nurses))))
    rand_seed = random.randint(1, 42)
    param_str = (
    f"randomization/randomseedshift={rand_seed}\n"
    f"randomization/lpseed=24\n"
    f"randomization/permutationseed=36"
)
    solver.SetSolverSpecificParametersAsString(param_str)

    # Solve
    status = solver.Solve()

    if status in [pywraplp.Solver.OPTIMAL, pywraplp.Solver.FEASIBLE]:
        # Build schedule with calendar days
        schedule = {}
        for n in nurses:
            schedule[n["id"]] = {}
            for d in range(total_days):
                date_str = all_dates[d].strftime('%Y-%m-%d')
                day_name = all_dates[d].strftime('%A')
                for s in shifts:
                    schedule[n["id"]][f"{day_name} {date_str} {s}"] = "Off"
        
        nurse_hours = {n["id"]: 0 for n in nurses}
        results_summary = []

        for d in range(total_days):
            current_date = all_dates[d]
            day_name = current_date.strftime('%A')
            day_result = {"Day": day_name, "Date": current_date.strftime('%Y-%m-%d')}
            for w in wards:
                for s in shifts:
                    s_idx = shift_indices[s]
                    assigned = sum(1 for n_idx in range(len(nurses)) 
                                   if x[n_idx, d, s_idx].solution_value() > 0.5 and 
                                   (w in nurses[n_idx].get("skills", []) or any(f"{w} {role}" in nurses[n_idx].get("skills", []) for role in ["Nurse", "Specialist", "Charge Nurse", "Nursing Officer", "Senior Staff Nurse"])))
                    day_result[f"{w}_{s}_predicted"] = int(week_demand.iloc[d][w])
                    day_result[f"{w}_{s}_assigned"] = assigned
                day_result[f"{w}_assigned_total"] = sum(day_result[f"{w}_{s}_assigned"] for s in shifts)
            results_summary.append(day_result)

            for n_idx, nurse in enumerate(nurses):
                for s_idx, s in enumerate(shifts):
                    if x[n_idx, d, s_idx].solution_value() > 0.5:
                        ward = next((w for w in wards if w in nurse.get("skills", []) or any(f"{w} {role}" in nurse.get("skills", []) for role in ["Nurse", "Specialist", "Charge Nurse", "Nursing Officer", "Senior Staff Nurse"])), None)
                        if ward:
                            date_str = all_dates[d].strftime('%Y-%m-%d')
                            day_name = all_dates[d].strftime('%A')
                            schedule[nurse["id"]][f"{day_name} {date_str} {s}"] = f"On Duty - {ward}"
                            nurse_hours[nurse["id"]] += SHIFT_HOURS

        # Convert to DataFrame with nurse names and IDs
        schedule_df = pd.DataFrame.from_dict(schedule, orient="index")
        
        # Create proper name mapping
        nurse_names = []
        for n in nurses:
            name = n.get("name", f"Unknown {n['id']}")
            nurse_names.append(name)
        
        schedule_df.index = nurse_names
        
        # Add nurse ID column and Name column
        schedule_df.insert(0, 'Nurse_ID', [n['id'] for n in nurses])
        schedule_df.insert(1, 'Name', nurse_names)

        return schedule_df, pd.DataFrame(results_summary), nurse_hours

    else:
        print("No optimal solution found.")
        return None, None, None

# ---------------- Main ----------------
if __name__ == "__main__":
    
    print("Loading models...")
    models = load_models()

    print("Loading dataset...")
    df = pd.read_csv(DATASET_CSV, parse_dates=["Date"], dayfirst=True)

    print("Loading nurses database...")
    with open(NURSE_JSON, "r", encoding="utf-8") as f:
        nurses = json.load(f)
    print(f"Loaded {len(nurses)} nurses from {NURSE_JSON}")

    print("Predicting next 7 days demand...")
    week_demand = predict_next_week(df, models, days=7)

    print("Scheduling nurses for next week...")
    schedule_df, summary_df, nurse_hours = schedule_nurses_optimized(week_demand, nurses, wards, shifts, weeks=1)

    if schedule_df is not None:
        # Save to Excel
        with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl") as writer:
            schedule_df.to_excel(writer, sheet_name="Schedule")
            summary_df.to_excel(writer, sheet_name="Summary", index=False)
            pd.DataFrame(list(nurse_hours.items()), columns=["Nurse ID", "Hours"]).to_excel(writer, sheet_name="Hours", index=False)

        # Apply coloring
        wb = load_workbook(OUTPUT_EXCEL)
        ws = wb["Schedule"]

        for row in ws.iter_rows(min_row=2, min_col=2):  # skip header and index col
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("On Duty"):
                    cell.fill = fill_on_duty
                elif cell.value == "Off":
                    cell.fill = fill_off

        wb.save(OUTPUT_EXCEL)
        print(f"Schedule saved to {OUTPUT_EXCEL} successfully. All nurses included!")
